# -*- coding: utf-8 -*-
import json
import uuid
import math

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse, Http404
import numpy as np
import pandas as pd

from sklearn import preprocessing, linear_model
from openpyxl import Workbook


from . import models

from .zhuanlu import PRO_BOF_HIS_ALLFIELDS
from QinggangManageSys.settings import MAIN_OUTFIT_BASE,MEDIA_ROOT,MEDIA_URL

from . import models
db_conn = models.BaseManage()
hello = "hello world"
from functools import partial
select = partial(db_conn.select, db_name='l2own')
execute = partial(db_conn.execute, db_name='l2own')
# def load(request):
#     if not request.user.is_authenticated():
#         return HttpResponseRedirect("/login")
#     contentVO={
#         'title':'数据分析工具',
#         'state':'success'
#     }
#     return render(request, 'data_import/analysis_tool.htm',contentVO)

def wushu_ana(df):
    """
    五数区间
    """
    Q1 = np.percentile(df,25)
    Q3 = np.percentile(df,75)
    L = 2*Q1 - Q3
    H = 2*Q3 - Q1
    ana_result={}
    ana_result['Q1'] = Q1
    ana_result['Q3'] = Q3
    ana_result['down'] = L
    ana_result['top'] = H
    return ana_result

def relation_cal(DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER, INDEPENDENT_VARIABLE_ATTRIBUTE_NUMBER):
    status = "fail"
    valid_num = 100
    keyno = '%s%s'%(DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER, INDEPENDENT_VARIABLE_ATTRIBUTE_NUMBER)
    filenames = {
        '32':"out_mid.txt",
        '31':"out_in.txt",
        '21':"mid_in.txt"
    }
    db_table_names = {
        '32':"relation_cof_output_middle",
        '31':"relation_cof_output_input",
        '21':"relation_cof_middle_input"
    }
    filename = filenames[keyno]
    db_table_name = db_table_names[keyno]

    A_data = []  # 输出字段
    A_isFiveAnalyse = dict()
    A_Range_Low = dict()
    A_Range_High = dict()

    B_data = []  # 控制字段
    B_isFiveAnalyse = dict()
    B_Range_High = dict()
    B_Range_Low = dict()

    if DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER ==3:
        sql = 'select DATA_ITEM_EN,IF_FIVENUMBERSUMMARY,NUMERICAL_LOWER_BOUND,NUMERICAL_UPPER_BOUND from qg_user.PRO_BOF_HIS_ALLSTRUCTURE WHERE (FIELD_ATTRIBUTE_NUMBER = %s OR FIELD_ATTRIBUTE_NUMBER = %s) AND IF_ANALYSE_TEMP = 1'% (DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER, DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER+1)
    else:
        sql = 'select DATA_ITEM_EN,IF_FIVENUMBERSUMMARY,NUMERICAL_LOWER_BOUND,NUMERICAL_UPPER_BOUND from qg_user.PRO_BOF_HIS_ALLSTRUCTURE WHERE FIELD_ATTRIBUTE_NUMBER = %s AND IF_ANALYSE_TEMP = 1'% DEPENDENT_VARIABLE_ATTRIBUTE_NUMBER
    print(sql)
    one = select(sql)
    if one == False:
        return status
    for row in one:
        A_data.append('%s' % row[0])
        A_isFiveAnalyse['%s' % row[0]] = '%s' % row[1]
        A_Range_Low['%s' % row[0]] = '%s' % row[2]
        A_Range_High['%s' % row[0]] = '%s' % row[3]
    sql = 'select DATA_ITEM_EN,IF_FIVENUMBERSUMMARY,NUMERICAL_LOWER_BOUND,NUMERICAL_UPPER_BOUND from qg_user.PRO_BOF_HIS_ALLSTRUCTURE WHERE FIELD_ATTRIBUTE_NUMBER = %s AND IF_ANALYSE_TEMP = 1'% INDEPENDENT_VARIABLE_ATTRIBUTE_NUMBER
    one = select(sql)
    if one == False:
        return status
    for row in one:
        B_data.append('%s' % row[0])
        B_isFiveAnalyse['%s' % row[0]] = '%s' % row[1]
        B_Range_Low['%s' % row[0]] = '%s' % row[2]
        B_Range_High['%s' % row[0]] = '%s' % row[3]

    #create table to save data.
    drop_sql = 'drop table QUERY.%s' % db_table_name
    create_sql = 'create table QUERY.%s(outputfield varchar(50),middlefield varchar(50),cof varchar(50))'% db_table_name
    execute(drop_sql)
    execute(create_sql)
    """
    iterate output and middields.
    """
    fout_om = open(filename, 'w+', encoding='utf-8')
    fout_error = open(db_table_name + '_error.txt','w+', encoding='utf-8')
    r_vlaue_array = []
    i = 0
    for A_liter in A_data:
        #add tag for A fields, just clean one time quarter number
        cleaningtag = True
        for B_liter in B_data:
            if A_liter != 'AS' and B_liter != 'AS':
                print(A_liter, B_liter)
                print(i)
                n = 0
                sx = 0
                sy = 0
                sxy = 0
                sx2 = 0
                sy2 = 0
                sql = 'select ' + A_liter + ',' + B_liter + ' from qg_user.PRO_BOF_HIS_ALLFIELDS '
                one = select(sql)
                if one == False:
                    return status
                abdf = pd.DataFrame(list(one))
                # 获取字段的上下数值范围
                A_Bound_Low = float(A_Range_Low[A_liter])
                A_Bound_High = float(A_Range_High[A_liter])
                B_Bound_Low = float(B_Range_Low[B_liter])
                B_Bound_High = float(B_Range_High[B_liter])
                print(A_Bound_Low, A_Bound_High, B_Bound_Low, B_Bound_High)
                # drop 0 and None values
                cols = list(abdf.columns)
                temp_df = abdf.copy()
                temp_df[cols[0]] = abdf[cols[0]].dropna().map(lambda x:float(x))
                # filter data by bound of low and high
                temp_df = temp_df[(temp_df[cols[0]] >= A_Bound_Low) & (temp_df[cols[0]] <=A_Bound_High)]
                # 获取相应字段的五数分析结果
                if cleaningtag:
                    if len(temp_df) == 0:
                        err_message =  'variable A %s has no valid data \n' % A_liter
                        print(err_message)
                        fout_error.write(err_message)
                        break
                    # 根据数据的上下限进行筛选
                    LHA = wushu_ana(temp_df.sort_values(by=cols[0])[cols[0]])
                    A_down = LHA['down']
                    A_top = LHA['top']
                    print(LHA,temp_df.sort_values(by=cols[0])[cols[0]].describe())
                    cleaningtag = False

                temp_df = abdf.copy()
                temp_df[cols[1]] = abdf[cols[1]].dropna().map(lambda x:float(x))
                temp_df = temp_df[(temp_df[cols[1]] >= B_Bound_Low) & (temp_df[cols[1]] <=B_Bound_High)]
                #加上数据的上下限
                if len(temp_df) == 0:
                    r_value = 0
                    fout_om.write('%s,%s,%s\n' % (A_liter, B_liter, r_value))
                    sql = 'insert into QUERY.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, A_liter, B_liter, r_value)
                    execute(sql)
                    err_message = 'variable B %s has no valid data \n' % B_liter
                    print(err_message)
                    fout_error.write(err_message)
                    i = i + 1
                    continue
                LHB = wushu_ana(temp_df.sort_values(by=cols[1])[cols[1]])
                B_down = LHB['down']
                B_top = LHB['top']
                print(LHB,temp_df.sort_values(by=cols[1])[cols[1]].describe())
                # 数据联合清洗，判断数据上下限
                try:
                    no_0_none_abdf = abdf.dropna(how='any')
                    no_0_none_abdf[cols[0]] = no_0_none_abdf[cols[0]].map(lambda x: float(x))
                    no_0_none_abdf[cols[1]] = no_0_none_abdf[cols[1]].map(lambda x: float(x))
                    no_0_none_abdf = no_0_none_abdf[(no_0_none_abdf[cols[0]] >= A_Bound_Low ) & (no_0_none_abdf[cols[1]] >= B_Bound_Low ) & (no_0_none_abdf[cols[0]] <= A_Bound_High ) & (no_0_none_abdf[cols[1]] <= B_Bound_High )]
                except:
                    r_value = 0
                    fout_om.write('%s,%s,%s\n' % (A_liter, B_liter, r_value))
                    sql = 'insert into QUERY.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, A_liter, B_liter, r_value)
                    execute(sql)
                    print(err_message)
                    fout_error.write(err_message)
                    continue

                if len(no_0_none_abdf) < valid_num:
                    r_value = 0
                    fout_om.write('%s,%s,%s\n' % (A_liter, B_liter, r_value))
                    sql = 'insert into QUERY.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, A_liter, B_liter, r_value)
                    execute(sql)
                    err_message = 'two vriables %s ,%s have not enough valid data \n' % (A_liter, B_liter)
                    print(err_message)
                    fout_error.write(err_message)
                    i = i + 1
                    continue

                print(A_down, A_top, B_down, B_top)

                """
                operate FiveAnalyse in responde to the structure.
                """

                if A_isFiveAnalyse[A_liter] == '1' and B_isFiveAnalyse[B_liter] == '1':
                    #print('A, B all need FiveAnalyse.')
                    no_0_none_abdf = no_0_none_abdf[(no_0_none_abdf[cols[0]] > A_down ) & (no_0_none_abdf[cols[0]] < A_top )][(no_0_none_abdf[cols[1]] > B_down ) & (no_0_none_abdf[cols[1]] < B_top )]
                elif A_isFiveAnalyse[A_liter] == '1' and B_isFiveAnalyse[B_liter] == '0':
                    #print('A needs FiveAnalyse.')
                    no_0_none_abdf = no_0_none_abdf[(no_0_none_abdf[cols[0]] > A_down ) & (no_0_none_abdf[cols[0]] < A_top )]
                elif A_isFiveAnalyse[A_liter] == '0' and B_isFiveAnalyse[B_liter] == '1':
                    #print('B needs FiveAnalyse.')
                    no_0_none_abdf = no_0_none_abdf[(no_0_none_abdf[cols[1]] > B_down ) & (no_0_none_abdf[cols[1]] < B_top )]
                elif A_isFiveAnalyse[A_liter] == '0' and B_isFiveAnalyse[B_liter] == '0':
                    print('Without FiveAnalyse.')

                if len(no_0_none_abdf) == 0:
                    r_value = 0
                    fout_om.write('%s,%s,%s\n' % (A_liter, B_liter, r_value))
                    err_message = '%s,%s no data after isFiveAnalyse compare \n' % (A_liter, B_liter)
                    print(err_message)
                    fout_error.write(err_message)
                    i = i + 1
                    continue
                """
                count A2 B2 AB
                """

                no_0_none_abdf['A2'] = no_0_none_abdf[cols[0]]*no_0_none_abdf[cols[0]]
                no_0_none_abdf['B2'] = no_0_none_abdf[cols[1]]*no_0_none_abdf[cols[1]]
                no_0_none_abdf['AB'] = no_0_none_abdf[cols[0]]*no_0_none_abdf[cols[1]]

                des_rs = dict(no_0_none_abdf.describe())
                keys = list(no_0_none_abdf.columns)

                n = int(des_rs[keys[0]]['count'])
                sx = n * des_rs[keys[0]]['mean']
                sy = n * des_rs[keys[1]]['mean']
                sx2 = n * des_rs[keys[2]]['mean']
                sy2 = n * des_rs[keys[3]]['mean']
                sxy = n * des_rs[keys[4]]['mean']

                """
                count r_value
                """
                r_value = 0

                denominator = float(math.sqrt(abs( n *sx2 - sx*sx )) * math.sqrt(abs( n * sy2 - sy*sy )))
                if  denominator != 0:
                    molecule = float(n * sxy - sx * sy )
                    r_value = molecule/denominator

                i = i + 1
                print(i)
                if i%20 == 0:
                    print(i)
                fout_om.write('%s,%s,%s\n' % (A_liter, B_liter, r_value))
                sql = 'insert into QUERY.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, A_liter, B_liter, r_value)
                execute(sql)
    fout_om.close()
    # fout_error.close()
    status = "success"
    return status

def relation_ana(request):
    if not request.user.is_authenticated():
            return HttpResponseRedirect("/login")
    contentVO={
        'title':'数据分析工具——关联性分析',
        'state':'success'
    }
    response_code = dict()
    """
    3 - output
    2 - control
    1 - input
    """
    entrance = [(3, 2), (3, 1), (2, 1)]
    for variables in entrance:
        x,y = variables
        key = '%s%s'%(x, y)
        response_code[key] = relation_cal(x,y)
    contentVO["response_code"] = response_code
    return HttpResponse(json.dumps(contentVO), content_type='application/json')

def data_cleaning(data):
    return data.fillna(0)

def linear_regression_model(data, isstd):
    cols = list(data.columns)
    length_cols = len(cols)
    features = cols[:-1]
    out = cols[-1]
    if isstd == 1:
        data_array = np.array(data)
        scale_data_array = preprocessing.scale(data_array)
        data_scale_df = pd.DataFrame(scale_data_array,columns=cols)

        X_scale = data_scale_df[features]
        Y_scale = data_scale_df[out]

    else:
        X_scale = data[features]
        Y_scale = data[out]

    regr = linear_model.LinearRegression()
    regr.fit(X_scale, Y_scale)

    return regr.coef_, regr.intercept_

def regression(output,selected_eles,isstd):
    """
    @param output 回归应变量 str
    @param selected_eles 回归自变量list
    @param isstd 是否进行标准化 1 标准化 0 不做标准化
    @rtn coef 与selected_eles对应的回归系数list
    @rtn intercep 回归方程截距
    @warning: 如果清洗后无数据，返回false，注意处理这种情况
    """
    fout_des = open("data_number.txt", 'a+', encoding='utf-8')
    sqlVO = {"db_name": 'l2own'}

    isFiveAnalyse = dict()
    bound_lows = dict()
    bound_highs = dict()

    sql = 'select DATA_ITEM_EN,IF_FIVENUMBERSUMMARY,NUMERICAL_LOWER_BOUND,NUMERICAL_UPPER_BOUND from qg_user.PRO_BOF_HIS_ALLSTRUCTURE WHERE  IF_ANALYSE_TEMP = 1'
    sqlVO["sql"] = sql
    rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)

    for row in rs:

        isFiveAnalyse['%s' % row[0]] = '%s' % row[1]
        bound_lows['%s' % row[0]] = '%s' % row[2]
        bound_highs['%s' % row[0]] = '%s' % row[3]

    allcolumns = selected_eles + [output]
    columns_coma = ",".join(allcolumns)
    sql = 'SELECT ' + ', '.join(selected_eles) + ', ' + output + ' from qg_user.PRO_BOF_HIS_ALLFIELDS'
    sqlVO["sql"] = sql
    rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)
    """
    移入平台时需要将tuple类型的result转化为list
    """
    alldf = pd.DataFrame(list(rs), columns=allcolumns)
    # data_ready = data_cleaning(pd.DataFrame(rs, columns=allcolumns))
    five_downs = dict()
    five_highs = dict()

    for col in allcolumns:
        temp_df = alldf.copy()
        temp_df[col] = temp_df[col].dropna().map(lambda x:float(x))

        # """
        # 可添加 用均值对原数据集空值的填充,若不需要则注释
        # """
        # mean=temp_df[col].describe().get('mean',0)
        # # print('mean:',mean)
        # alldf[col] = alldf[col].fillna(mean)
        # """
        # end fill nan with mean
        # """

        #filter data by bound of low and high
        bound_low = float(bound_lows.get(col,-999999999999))
        bound_high = float(bound_highs.get(col,999999999999))
        temp_df = temp_df[(temp_df[col] >= bound_low) & ( temp_df[col] <= bound_high )]
        if isFiveAnalyse.get(col,'0') == '1':
            LH = wushu_ana(temp_df.sort_values(by=col)[col])
            five_downs['%s' % col ] = LH['down']
            five_highs['%s' % col ] = LH['top']
            print(col, five_downs[col], five_highs[col])

    """
    根据上下限填充零值
    """
    for col in allcolumns:
        bound_low = float(bound_lows.get(col,-999999999999))
        bound_high = float(bound_highs.get(col,999999999999))
        if 0 >= bound_low and 0 <= bound_high:
            alldf[col] = alldf[col].fillna(0)

    alldf_temp = alldf.copy()
    # fout_des.write("%s before drop : %s\n" % (columns_coma,str(len(alldf))))
    before_drop_num = len(alldf)
    alldf = alldf.dropna(how='any')
    after_drop_num = len(alldf)
    fout_des.write("%s  nan rate : %s\n" % (columns_coma,str(before_drop_num/after_drop_num)))
    for col in allcolumns:
        temp_col_s = alldf[col]
        temp_col_s = temp_col_s.dropna()
        after_drop_temp_col = len(temp_col_s)
        fout_des.write("%s  nan rate : %s\n" % (col,str(after_drop_temp_col/after_drop_num)))

    # 有字段的类型为Object
    for col in allcolumns:
        alldf[col] = alldf[col].map(lambda x:float(x))

    # 根据各因素上限联合筛选数据
    value_bound_tag = True
    for col in allcolumns:
        bound_low = float(bound_lows.get(col,-999999999999))
        bound_high = float(bound_highs.get(col,999999999999))
        alldf = alldf[(alldf[col] >= bound_low) & (alldf[col] <= bound_high)]
        if len(alldf) == 0:
            print("no data after bound")
            value_bound_tag = False
            break

    if  not value_bound_tag:
        return False
    # 五值分析
    five_tag = True
    for col in allcolumns:
        if isFiveAnalyse.get(col,'0') == '1':
            five_down = five_downs.get(col,-999999999999)
            five_high = five_highs.get(col,999999999999)
            alldf = alldf[( alldf[col] >= five_down ) & ( alldf[col] <= five_high )]
            if len(alldf) == 0:
                five_tag = False
                print("no data after five")
                break
    if not five_tag:
        return False

    coef, intercept = linear_regression_model(alldf, isstd)
    coef = list(map(lambda x: str(x), coef))

    # save result to database
    """
    for i in range(len(selected_eles)):
        sql = 'insert into QUERY.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, output, selected_eles[i], coef[i])
        sqlVO['sql'] = sql
        models.BaseManage().direct_execute_query_sqlVO(sqlVO)
    sql = 'insert into QUERY.%s values(\'%s\', \'BIAS\', \'%s\')'%(db_table_name, output, str(intercept))
    sqlVO['sql'] = sql
    models.BaseManage().direct_execute_query_sqlVO(sqlVO)
    """
    fout_des.close()
    return coef, intercept



def regression_ana(result):
    pass


def report(request):
    if not request.user.is_authenticated():
            return HttpResponseRedirect("/login")
    contentVO={
        'title':'数据分析工具——生成分析结果报表',
        'state':'success'
    }
    db_table_names = {
        '321':"relation_cof_output_middle",
        '311':"relation_cof_output_input",
        '211':"relation_cof_middle_input",
        '322':"regression_cof_output_middle",
        '312':"regression_cof_output_input",
        '212':"regression_cof_middle_input",
    }
    sheetnames = {
        "relation_cof_output_middle":"相关系数_输出-控制",
        "relation_cof_output_input":"相关系数_输出-输入",
        "relation_cof_middle_input":"相关系数_控制-输入",
        "regression_cof_output_middle":"回归系数_输出-控制",
        "regression_cof_output_input":"回归系数_输出-输入",
        "regression_cof_middle_input":"回归系数_控制-输入",
    }
    entrance = [(3,2,1) ,(3,1,1), (2,1,1), (3,2,2) ,(3,1,2), (2,1,2)]


    filename = '相关性_回归分析汇总%s.xlsx' % str(uuid.uuid1())
    save_filename = MEDIA_ROOT + '/' + filename
    print(save_filename)
    contentVO['filepath'] = MEDIA_URL + filename

    sqlVO = dict()
    sqlVO["db_name"] = "l2own"

    wb = Workbook()
    i = 0

    for tags in entrance:
        tablekey = '%s%s%s' % tags
        table_name = db_table_names[tablekey]

        sql = 'SELECT * FROM QUERY.%s' % table_name.upper()
        sqlVO["sql"] = sql

        rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)
        if len(rs) == 0:
            continue

        if i == 0:
            ws = wb.active
            ws.title = sheetnames[table_name]
            i += 1
        else:
            ws = wb.create_sheet(sheetnames[table_name])

        rs_df = pd.DataFrame(list(rs))

        #sorted by column 1 and 3
        cols = list(rs_df.columns)
        rs_df['abs'] = rs_df[cols[2]].map(lambda x: abs(float(x)))
        # fout_sort = open('sort.txt', 'w+',encoding='utf-8')
        rs_df = rs_df.sort_values(by=[cols[0],'abs'],ascending=False)
        for ind in rs_df.index:
            temp = tuple(rs_df.loc[ind,cols])
            keyA, keyB, value = temp
            keyB = keyB.strip()
            if keyB == 'BIAS':
                nameB = "偏离-截距"
            else:
                nameB = PRO_BOF_HIS_ALLFIELDS[keyB]
            ws.append((PRO_BOF_HIS_ALLFIELDS[keyA.strip()], nameB, value))
            # fout_sort.write('%s,%s,%s\n'% tuple(rs_df.loc[ind,cols]))
        # fout_sort.close()
    print(save_filename)
    wb.save(save_filename)
    return HttpResponse(json.dumps(contentVO), content_type='application/json')


'''
转炉数据提取
'''

'''
按时执行定时任务
'''
def batch_relation_ana():
    response_code = dict()
    """
    3 - output
    2 - control
    1 - input
    """
    entrance = [(3, 2), (3, 1), (2, 1)]
    for variables in entrance:
        x,y = variables
        key = '%s%s'%(x, y)
        response_code[key] = relation_cal(x,y)
    return response_code
